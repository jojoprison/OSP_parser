from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()

print(wb)
print(wb.active)

ws1 = wb.create_sheet("MySheet")
ws2 = wb.create_sheet("MySheet2", 0)
ws3 = wb.create_sheet("My", -1)

print(ws1.title)

ws1.title = "SOAP"
print(ws1.title)

ws1.sheet_properties.tabColor = "1072BA"

print(wb.sheetnames)

wb.remove(ws2)
wb.remove(ws3)

print(wb.sheetnames)

del wb['Sheet']

print(wb.sheetnames)

dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

for row in range(1, 40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title="Pi")

ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 40):
    for col in range(27, 54):
        ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))

print(ws3['AA10'].value)

wb.save(filename=dest_filename)
